import pandas as pd
import openpyxl as xl

wb = xl.Workbook()
sheet = wb.active
sheet.title = '수험데이터 변환'

col_names = ['수험번호', '이름']
for seq, name in enumerate(col_names):
    sheet.cell(row=1, column=seq+1, value=name)


xlsx = pd.read_excel('/Users/tyflow/Downloads/scoreData.xlsx')
print(xlsx)

temp = list(xlsx['이름'])

row_no = 2
for n, row in enumerate(temp):
    tempRow = row.split('_')
    for seq, value in enumerate(tempRow):
        sheet.cell(row=row_no + n, column=seq + 1, value=value)

wb.save('/Users/tyflow/Downloads/result.xlsx')
